"""
Excel Export - FAiLLON Development Pro Forma

Generates Excel workbooks with:
- Live formulas (no hardcoded values)
- Parcel-to-GFA analysis with city-specific FAR
- Floor and unit density calculations
- Scenario analysis (N, N±10, N±20 units)
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Styles
FILL_INPUT = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_CALC = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_SECTION = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_TITLE = Font(bold=True, size=14)
FONT_BOLD = Font(bold=True, size=10)
FONT_NORMAL = Font(size=10)
FONT_INPUT = Font(bold=True, size=10, color="000080")
FONT_SMALL = Font(size=9, italic=True, color="666666")

BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _val(section: dict, key: str, default=None):
    if not isinstance(section, dict):
        return default
    field = section.get(key)
    if field is None:
        return default
    if isinstance(field, dict):
        v = field.get("value")
        return v if v is not None else default
    return field


def _col(n):
    return get_column_letter(n)


def _ref(sheet, row, col):
    return f"'{sheet}'!{_col(col)}{row}"


# ═══════════════════════════════════════════════════════════════════════════════
# INPUTS SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def _build_inputs(wb, data):
    ws = wb.create_sheet("Inputs", 0)
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 45
    
    pf = data.get("pro_forma", {})
    summary = pf.get("project_summary", {})
    revenue = pf.get("revenue_assumptions", {})
    costs = pf.get("cost_assumptions", {})
    financing = pf.get("financing_assumptions", {})
    returns = pf.get("return_metrics", {})
    
    refs = {}
    row = 1
    
    # Title
    ws.cell(row, 1, "FAiLLON DEVELOPMENT PRO FORMA").font = FONT_TITLE
    row += 1
    ws.cell(row, 1, "Yellow = editable inputs | Green = calculated").font = FONT_SMALL
    row += 2
    
    # Headers
    for c, h in enumerate(["Parameter", "Value", "Unit", "Notes"], 1):
        cell = ws.cell(row, c, h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.border = BORDER
    row += 1
    
    def section(title):
        nonlocal row
        cell = ws.cell(row, 1, title)
        cell.font = FONT_BOLD
        cell.fill = FILL_SECTION
        for c in range(1, 5):
            ws.cell(row, c).fill = FILL_SECTION
            ws.cell(row, c).border = BORDER
        row += 1
    
    def inp(key, label, val, unit, note):
        nonlocal row
        ws.cell(row, 1, label).font = FONT_NORMAL
        cell = ws.cell(row, 2, val if val is not None else 0)
        cell.font = FONT_INPUT
        cell.fill = FILL_INPUT
        cell.border = BORDER
        if "%" in unit:
            cell.number_format = '0.00'
        elif "$" in unit or "cost" in label.lower():
            cell.number_format = '#,##0'
        ws.cell(row, 3, unit).font = FONT_NORMAL
        ws.cell(row, 4, note).font = FONT_SMALL
        refs[key] = row  # Store row number
        row += 1
    
    # PROJECT
    section("PROJECT BASICS")
    inp("deal_name", "Deal Name", _val(summary, "deal_name", "Development"), "", "")
    inp("market", "Market", _val(summary, "market", "Charlotte"), "", "Boston/Charlotte/Nashville")
    inp("program", "Program Type", _val(summary, "program_type", "multifamily"), "", "")
    
    row += 1
    section("PARCEL & DENSITY")
    inp("parcel_sf", "Parcel Size (SF)", _val(summary, "parcel_sf", 43560), "SF", "1 acre = 43,560 SF")
    inp("parcel_acres", "Parcel Size (Acres)", _val(summary, "acreage", 1.0), "acres", "= SF / 43,560")
    inp("far", "Floor Area Ratio (FAR)", _val(summary, "far", 3.0), "x", "City zoning - Boston 4-8, Charlotte 2-4")
    inp("efficiency", "Building Efficiency", _val(summary, "efficiency_pct", 85), "%", "Rentable / GFA")
    inp("floors", "Number of Floors", _val(summary, "floors", 5), "floors", "Based on zoning/FAR")
    
    row += 1
    section("UNIT MIX")
    inp("unit_count", "Total Units", _val(summary, "unit_count", 100), "units", "")
    inp("avg_unit_sf", "Avg Unit Size", _val(summary, "avg_unit_sf", 950), "SF", "")
    inp("total_gfa", "Total GFA", _val(summary, "total_gfa_sf", 100000), "SF", "")
    inp("keys", "Hotel Keys", _val(summary, "total_keys", 0), "keys", "If hotel")
    
    row += 1
    section("REVENUE")
    inp("rent_psf", "Monthly Rent", _val(revenue, "rent_psf_monthly", 2.50), "$/SF/mo", "")
    inp("occupancy", "Stabilized Occupancy", _val(revenue, "stabilized_occupancy_pct", 95), "%", "")
    inp("other_income", "Other Income", _val(revenue, "other_income_pct_egi", 5), "% of rent", "")
    inp("rent_growth", "Annual Rent Growth", _val(revenue, "annual_rent_growth_pct", 3), "%", "")
    inp("opex_ratio", "Operating Expense Ratio", _val(revenue, "opex_ratio_pct", 35), "%", "")
    
    row += 1
    section("COSTS")
    inp("land_total", "Land Cost (Total)", _val(costs, "land_cost_total", 0), "$", "Or use per acre")
    inp("land_per_acre", "Land Cost per Acre", _val(costs, "land_cost_per_acre", 500000), "$/acre", "")
    inp("hard_psf", "Hard Cost", _val(costs, "hard_cost_psf", 250), "$/SF", "Construction")
    inp("soft_pct", "Soft Cost %", _val(costs, "soft_cost_pct_of_hard", 25), "%", "Of hard")
    inp("contingency", "Contingency %", _val(costs, "contingency_pct", 5), "%", "")
    inp("dev_fee", "Developer Fee %", _val(costs, "developer_fee_pct", 4), "%", "")
    
    row += 1
    section("FINANCING")
    inp("ltc", "Loan to Cost", _val(financing, "construction_loan_ltc_pct", 65), "%", "")
    inp("rate", "Interest Rate", _val(financing, "construction_loan_rate_pct", 7.5), "%", "")
    inp("lp_pct", "LP Equity %", _val(financing, "lp_equity_pct", 90), "%", "")
    
    row += 1
    section("EXIT")
    inp("exit_cap", "Exit Cap Rate", _val(returns, "exit_cap_rate_pct", 5.25), "%", "")
    inp("hold", "Hold Period", _val(returns, "exit_year", 5), "years", "")
    inp("sale_cost", "Sale Costs", _val(returns, "sale_costs_pct", 2), "%", "")
    
    return refs


# ═══════════════════════════════════════════════════════════════════════════════
# PARCEL ANALYSIS SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def _build_parcel_analysis(wb, refs):
    ws = wb.create_sheet("Parcel Analysis")
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 45
    
    row = 1
    ws.cell(row, 1, "PARCEL TO GFA ANALYSIS").font = FONT_TITLE
    row += 2
    
    # Headers
    for c, h in enumerate(["Metric", "Value", "Formula/Notes"], 1):
        cell = ws.cell(row, c, h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.border = BORDER
    row += 1
    
    def calc(label, formula, note, fmt='#,##0'):
        nonlocal row
        ws.cell(row, 1, label).font = FONT_NORMAL
        cell = ws.cell(row, 2)
        cell.value = formula
        cell.number_format = fmt
        cell.fill = FILL_CALC
        cell.border = BORDER
        ws.cell(row, 3, note).font = FONT_SMALL
        row += 1
        return row - 1
    
    def section(title):
        nonlocal row
        cell = ws.cell(row, 1, title)
        cell.font = FONT_BOLD
        cell.fill = FILL_SECTION
        for c in range(1, 4):
            ws.cell(row, c).fill = FILL_SECTION
            ws.cell(row, c).border = BORDER
        row += 1
    
    # Parcel inputs
    section("PARCEL INPUTS")
    r_sf = calc("Parcel SF", f"=Inputs!B{refs['parcel_sf']}", "From inputs")
    r_acres = calc("Parcel Acres", f"=Inputs!B{refs['parcel_sf']}/43560", "SF / 43,560", '0.00')
    
    row += 1
    section("ZONING & DENSITY")
    r_far = calc("FAR (Floor Area Ratio)", f"=Inputs!B{refs['far']}", "City zoning allowance", '0.0')
    r_max_gfa = calc("Maximum Buildable GFA", f"=B{r_sf}*B{r_far}", "Parcel SF × FAR")
    
    row += 1
    section("BUILDING CONFIGURATION")
    r_floors = calc("Number of Floors", f"=Inputs!B{refs['floors']}", "Based on zoning/height limits")
    r_floorplate = calc("Typical Floorplate", f"=B{r_max_gfa}/B{r_floors}", "GFA / Floors")
    r_eff = calc("Building Efficiency", f"=Inputs!B{refs['efficiency']}/100", "Rentable / Gross", '0%')
    r_rentable = calc("Total Rentable SF", f"=B{r_max_gfa}*B{r_eff}", "GFA × Efficiency")
    
    row += 1
    section("UNIT ANALYSIS")
    r_avg_sf = calc("Avg Unit Size (Input)", f"=Inputs!B{refs['avg_unit_sf']}", "From inputs")
    r_eff_units = calc("Effective Unit Count", f"=ROUND(B{r_rentable}/B{r_avg_sf},0)", "Rentable SF / Avg Unit SF")
    r_units_floor = calc("Units per Floor", f"=ROUND(B{r_eff_units}/B{r_floors},0)", "Total Units / Floors")
    
    row += 1
    section("DENSITY METRICS")
    calc("Units per Acre", f"=B{r_eff_units}/B{r_acres}", "Units / Acres", '0.0')
    calc("FAR Used", f"=Inputs!B{refs['total_gfa']}/B{r_sf}", "Actual GFA / Parcel", '0.00')
    calc("GFA per Unit", f"=Inputs!B{refs['total_gfa']}/Inputs!B{refs['unit_count']}", "Gross SF / Units")
    
    row += 2
    ws.cell(row, 1, "CITY FAR GUIDELINES:").font = FONT_BOLD
    row += 1
    ws.cell(row, 1, "• Boston: FAR 4-8 (high density, transit-oriented)").font = FONT_SMALL
    row += 1
    ws.cell(row, 1, "• Charlotte: FAR 2-4 (mid-rise suburban/urban)").font = FONT_SMALL
    row += 1
    ws.cell(row, 1, "• Nashville: FAR 2-4 (similar to Charlotte)").font = FONT_SMALL


# ═══════════════════════════════════════════════════════════════════════════════
# CALCULATIONS SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def _build_calculations(wb, refs):
    ws = wb.create_sheet("Calculations")
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 45
    
    row = 1
    ws.cell(row, 1, "FINANCIAL CALCULATIONS").font = FONT_TITLE
    row += 2
    
    for c, h in enumerate(["Item", "Value", "Formula"], 1):
        cell = ws.cell(row, c, h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.border = BORDER
    row += 1
    
    def calc(label, formula, note, fmt='$#,##0'):
        nonlocal row
        ws.cell(row, 1, label).font = FONT_NORMAL
        cell = ws.cell(row, 2)
        cell.value = formula
        cell.number_format = fmt
        cell.fill = FILL_CALC
        cell.border = BORDER
        ws.cell(row, 3, note).font = FONT_SMALL
        row += 1
        return row - 1
    
    def section(title):
        nonlocal row
        cell = ws.cell(row, 1, title)
        cell.font = FONT_BOLD
        cell.fill = FILL_SECTION
        for c in range(1, 4):
            ws.cell(row, c).fill = FILL_SECTION
            ws.cell(row, c).border = BORDER
        row += 1
    
    # Revenue
    section("REVENUE")
    r_rent_sf = calc("Rentable SF", f"=Inputs!B{refs['unit_count']}*Inputs!B{refs['avg_unit_sf']}", "Units × Avg SF", '#,##0')
    r_gpr = calc("Gross Potential Rent", f"=B{r_rent_sf}*Inputs!B{refs['rent_psf']}*12", "Rentable × Rent × 12")
    r_other = calc("Other Income", f"=B{r_gpr}*Inputs!B{refs['other_income']}/100", "GPR × Other %")
    r_gpi = calc("Gross Potential Income", f"=B{r_gpr}+B{r_other}", "Rent + Other")
    r_vacancy = calc("Vacancy Loss", f"=B{r_gpi}*(1-Inputs!B{refs['occupancy']}/100)", "GPI × (1-Occ)")
    r_egi = calc("Effective Gross Income", f"=B{r_gpi}-B{r_vacancy}", "GPI - Vacancy")
    r_opex = calc("Operating Expenses", f"=B{r_egi}*Inputs!B{refs['opex_ratio']}/100", "EGI × OpEx%")
    r_noi = calc("Net Operating Income", f"=B{r_egi}-B{r_opex}", "EGI - OpEx")
    
    row += 1
    section("COSTS")
    r_land = calc("Land Cost", f"=IF(Inputs!B{refs['land_total']}>0,Inputs!B{refs['land_total']},Inputs!B{refs['land_per_acre']}*Inputs!B{refs['parcel_acres']})", "Total or Per Acre × Acres")
    r_hard = calc("Hard Costs", f"=Inputs!B{refs['total_gfa']}*Inputs!B{refs['hard_psf']}", "GFA × $/SF")
    r_soft = calc("Soft Costs", f"=B{r_hard}*Inputs!B{refs['soft_pct']}/100", "Hard × Soft%")
    r_cont = calc("Contingency", f"=(B{r_hard}+B{r_soft})*Inputs!B{refs['contingency']}/100", "(H+S) × Cont%")
    r_sub = calc("Subtotal", f"=B{r_land}+B{r_hard}+B{r_soft}+B{r_cont}", "Land+Hard+Soft+Cont")
    r_fee = calc("Developer Fee", f"=B{r_sub}*Inputs!B{refs['dev_fee']}/100", "Sub × Fee%")
    r_tdc = calc("Total Development Cost", f"=B{r_sub}+B{r_fee}", "Subtotal + Fee")
    ws.cell(row-1, 2).font = FONT_BOLD
    
    row += 1
    section("FINANCING")
    r_loan = calc("Loan Amount", f"=B{r_tdc}*Inputs!B{refs['ltc']}/100", "TDC × LTC%")
    r_equity = calc("Total Equity", f"=B{r_tdc}-B{r_loan}", "TDC - Loan")
    r_lp = calc("LP Equity", f"=B{r_equity}*Inputs!B{refs['lp_pct']}/100", "Equity × LP%")
    r_gp = calc("GP Equity", f"=B{r_equity}-B{r_lp}", "Equity - LP")
    
    row += 1
    section("RETURNS")
    r_exit_val = calc("Exit Value (Gross)", f"=B{r_noi}/(Inputs!B{refs['exit_cap']}/100)", "NOI / Cap Rate")
    r_sale_cost = calc("Sale Costs", f"=B{r_exit_val}*Inputs!B{refs['sale_cost']}/100", "Gross × Sale%")
    r_net_exit = calc("Exit Value (Net)", f"=B{r_exit_val}-B{r_sale_cost}", "Gross - Sale")
    r_profit = calc("Total Profit", f"=B{r_net_exit}-B{r_tdc}", "Net Exit - TDC")
    r_poc = calc("Profit on Cost", f"=B{r_profit}/B{r_tdc}", "Profit / TDC", '0.0%')
    calc("Cost per Unit", f"=B{r_tdc}/Inputs!B{refs['unit_count']}", "TDC / Units")
    calc("NOI Yield on Cost", f"=B{r_noi}/B{r_tdc}", "NOI / TDC", '0.0%')
    
    return r_tdc, r_noi


# ═══════════════════════════════════════════════════════════════════════════════
# UNIT SCENARIOS
# ═══════════════════════════════════════════════════════════════════════════════

def _build_scenarios(wb, refs, data):
    ws = wb.create_sheet("Unit Scenarios")
    for c, w in enumerate([25, 16, 16, 16, 16, 16], 1):
        ws.column_dimensions[_col(c)].width = w
    
    pf = data.get("pro_forma", {})
    base = _val(pf.get("project_summary", {}), "unit_count", 100) or 100
    
    row = 1
    ws.cell(row, 1, "UNIT MIX SCENARIO ANALYSIS").font = FONT_TITLE
    row += 1
    ws.cell(row, 1, "Same GFA, different unit counts = different avg sizes").font = FONT_SMALL
    row += 2
    
    scenarios = [("N-20", -20), ("N-10", -10), ("Base", 0), ("N+10", 10), ("N+20", 20)]
    
    # Headers
    for c, (lbl, _) in enumerate(scenarios, 2):
        cell = ws.cell(row, c, lbl)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    ws.cell(row, 1, "Metric").font = FONT_HEADER
    ws.cell(row, 1).fill = FILL_HEADER
    ws.cell(row, 1).border = BORDER
    row += 1
    
    def metric(label, formulas, fmt='#,##0'):
        nonlocal row
        ws.cell(row, 1, label).font = FONT_NORMAL
        ws.cell(row, 1).border = BORDER
        for c, f in enumerate(formulas, 2):
            cell = ws.cell(row, c)
            cell.value = f
            cell.number_format = fmt
            cell.border = BORDER
            cell.fill = FILL_CALC
        row += 1
        return row - 1
    
    # Unit counts
    unit_formulas = [f"=Inputs!B{refs['unit_count']}+{d}" for _, d in scenarios]
    r_units = metric("Unit Count", unit_formulas)
    
    # Avg unit size (fixed GFA)
    avg_formulas = [f"=Inputs!B{refs['total_gfa']}/{_col(c)}{r_units}" for c in range(2, 7)]
    r_avg = metric("Avg Unit SF", avg_formulas)
    
    # Rentable
    rent_formulas = [f"={_col(c)}{r_units}*{_col(c)}{r_avg}" for c in range(2, 7)]
    r_rent = metric("Rentable SF", rent_formulas)
    
    # Annual rent
    ann_formulas = [f"={_col(c)}{r_rent}*Inputs!B{refs['rent_psf']}*12" for c in range(2, 7)]
    r_ann = metric("Annual Rent", ann_formulas, '$#,##0')
    
    # EGI
    egi_formulas = [f"={_col(c)}{r_ann}*(1+Inputs!B{refs['other_income']}/100)*(Inputs!B{refs['occupancy']}/100)" for c in range(2, 7)]
    r_egi = metric("EGI", egi_formulas, '$#,##0')
    
    # NOI
    noi_formulas = [f"={_col(c)}{r_egi}*(1-Inputs!B{refs['opex_ratio']}/100)" for c in range(2, 7)]
    r_noi = metric("NOI", noi_formulas, '$#,##0')
    
    row += 1
    ws.cell(row, 1, "COSTS (Fixed GFA)").font = FONT_BOLD
    row += 1
    
    # TDC (same for all)
    tdc_f = f"=(IF(Inputs!B{refs['land_total']}>0,Inputs!B{refs['land_total']},Inputs!B{refs['land_per_acre']}*Inputs!B{refs['parcel_acres']})+Inputs!B{refs['total_gfa']}*Inputs!B{refs['hard_psf']}*(1+Inputs!B{refs['soft_pct']}/100)*(1+Inputs!B{refs['contingency']}/100))*(1+Inputs!B{refs['dev_fee']}/100)"
    r_tdc = metric("Total Dev Cost", [tdc_f]*5, '$#,##0')
    
    # Cost per unit
    cpu_formulas = [f"={_col(c)}{r_tdc}/{_col(c)}{r_units}" for c in range(2, 7)]
    metric("Cost per Unit", cpu_formulas, '$#,##0')
    
    row += 1
    ws.cell(row, 1, "RETURNS").font = FONT_BOLD
    row += 1
    
    # Exit value
    exit_formulas = [f"=({_col(c)}{r_noi}/(Inputs!B{refs['exit_cap']}/100))*(1-Inputs!B{refs['sale_cost']}/100)" for c in range(2, 7)]
    r_exit = metric("Exit Value (Net)", exit_formulas, '$#,##0')
    
    # Profit
    profit_formulas = [f"={_col(c)}{r_exit}-{_col(c)}{r_tdc}" for c in range(2, 7)]
    r_profit = metric("Profit", profit_formulas, '$#,##0')
    
    # PoC
    poc_formulas = [f"={_col(c)}{r_profit}/{_col(c)}{r_tdc}" for c in range(2, 7)]
    metric("Profit on Cost", poc_formulas, '0.0%')


# ═══════════════════════════════════════════════════════════════════════════════
# SENSITIVITY
# ═══════════════════════════════════════════════════════════════════════════════

def _build_sensitivity(wb, data):
    ws = wb.create_sheet("Sensitivity")
    for c, w in enumerate([18, 14, 14, 14], 1):
        ws.column_dimensions[_col(c)].width = w
    
    sens = data.get("sensitivity", {})
    
    row = 1
    ws.cell(row, 1, "SENSITIVITY ANALYSIS").font = FONT_TITLE
    row += 2
    
    cols = sens.get("cols", ["-10%", "Base", "+10%"])
    rows_labels = sens.get("rows", ["4.75%", "5.25%", "5.75%"])
    values = sens.get("values", [[15, 12, 9], [12, 10, 7], [9, 7, 5]])
    colors = sens.get("colors", [["green", "green", "yellow"], ["green", "yellow", "red"], ["yellow", "red", "red"]])
    
    ws.cell(row, 1, "Exit Cap \\ Cost").font = FONT_BOLD
    for c, h in enumerate(cols, 2):
        cell = ws.cell(row, c, h)
        cell.font = FONT_BOLD
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    row += 1
    
    color_map = {"green": FILL_GREEN, "yellow": FILL_YELLOW, "red": FILL_RED}
    
    for i, r_label in enumerate(rows_labels):
        cell = ws.cell(row, 1, r_label)
        cell.font = FONT_BOLD
        cell.border = BORDER
        for j, v in enumerate(values[i] if i < len(values) else [0]*3):
            cell = ws.cell(row, j+2, v)
            cell.number_format = '0.0"%"'
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            col = colors[i][j] if i < len(colors) and j < len(colors[i]) else "yellow"
            cell.fill = color_map.get(col, FILL_YELLOW)
        row += 1


# ═══════════════════════════════════════════════════════════════════════════════
# SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

def _build_summary(wb, refs, data):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    
    pf = data.get("pro_forma", {})
    name = _val(pf.get("project_summary", {}), "deal_name", "Pro Forma") or "Pro Forma"
    
    row = 1
    ws.cell(row, 1, str(name)).font = FONT_TITLE
    row += 2
    
    metrics = [
        ("Market", f"=Inputs!B{refs['market']}", ""),
        ("Parcel (Acres)", f"=Inputs!B{refs['parcel_acres']}", "0.00"),
        ("Total GFA", f"=Inputs!B{refs['total_gfa']}", "#,##0"),
        ("Unit Count", f"=Inputs!B{refs['unit_count']}", "#,##0"),
        ("Exit Cap", f"=Inputs!B{refs['exit_cap']}", "0.00\"%\""),
    ]
    
    for label, formula, fmt in metrics:
        ws.cell(row, 1, label).font = FONT_NORMAL
        cell = ws.cell(row, 2)
        cell.value = formula
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt
        row += 1


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def export_pro_forma(data: dict, deal_name: str = "Pro Forma") -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    
    refs = _build_inputs(wb, data)
    _build_parcel_analysis(wb, refs)
    _build_calculations(wb, refs)
    _build_scenarios(wb, refs, data)
    _build_sensitivity(wb, data)
    _build_summary(wb, refs, data)
    
    wb.active = wb["Inputs"]
    wb.calculation.calcMode = "auto"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_suggested_filename(data: dict) -> str:
    pf = data.get("pro_forma", {})
    name = _val(pf.get("project_summary", {}), "deal_name", "Pro_Forma") or "Pro_Forma"
    return f"{str(name).replace(' ', '_').replace('/', '-')}.xlsx"
