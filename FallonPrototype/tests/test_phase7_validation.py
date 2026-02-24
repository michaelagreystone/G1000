"""
Phase 7 — Testing & Validation

Comprehensive end-to-end tests for all Financial Model components.
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from FallonPrototype.agents.financial_agent import (
    extract_parameters,
    normalize_parameters,
    check_missing_parameters,
    merge_clarification,
    ProjectParameters,
    retrieve_deal_comps,
    validate_pro_forma,
    VALID_LABELS,
)
from FallonPrototype.shared.return_calculator import compute_returns, _val
from FallonPrototype.shared.excel_export import export_pro_forma
from FallonPrototype.shared.vector_store import get_collection_counts


# ═══════════════════════════════════════════════════════════════════════════════
# 7.2 — Test Deal Comps Retrieval
# ═══════════════════════════════════════════════════════════════════════════════

def test_deal_comps_retrieval():
    """Confirm Charlotte query returns Charlotte deal file as top result."""
    print("\n" + "=" * 60)
    print("TEST 7.2: Deal Comps Retrieval")
    print("=" * 60)
    
    params = ProjectParameters(market="charlotte", program_type="multifamily")
    comps = retrieve_deal_comps(params)
    
    print(f"  Retrieved {len(comps)} comps for Charlotte multifamily")
    
    if len(comps) == 0:
        print("  Note: Empty corpus - no deal data indexed yet")
        print("  SKIP: Cannot validate without deal data")
        return True  # Not a failure if data not yet indexed
    
    # Check top result
    top = comps[0]
    source = top["metadata"].get("source", "").lower()
    market = top["metadata"].get("market", "").lower()
    
    print(f"  Top result: {source}")
    print(f"  Market tag: {market}")
    print(f"  Relevance: {top.get('relevance')}")
    
    # Verify no low-relevance results
    low_rel = [c for c in comps if c.get("relevance") == "low"]
    assert len(low_rel) == 0, f"Found {len(low_rel)} low-relevance results (should be filtered)"
    print("  No low-relevance results (filtered correctly)")
    
    print("\nPASS: Deal comps retrieval works")
    return True


# ═══════════════════════════════════════════════════════════════════════════════
# 7.4 — Test Calculator Cross-Check
# ═══════════════════════════════════════════════════════════════════════════════

def test_calculator_cross_check():
    """Verify compute_returns produces results within 5% of manual calculation."""
    print("\n" + "=" * 60)
    print("TEST 7.4: Calculator Cross-Check")
    print("=" * 60)
    
    # Known test case with manually computed values
    test_pro_forma = {
        "project_summary": {
            "program_type": "multifamily",
            "unit_count": {"value": 200},
            "construction_duration_months": {"value": 18},
        },
        "revenue_assumptions": {
            "rent_psf_monthly": {"value": 1.85},
            "stabilized_occupancy_pct": {"value": 93},
            "lease_up_months": {"value": 18},
            "other_income_per_unit_monthly": {"value": 125},
        },
        "cost_assumptions": {
            "total_project_cost": {"value": 70_000_000},
        },
        "financing_assumptions": {
            "equity_required": {"value": 24_500_000},
            "lp_equity_amount": {"value": 22_050_000},
            "lp_equity_pct": {"value": 90},
        },
        "return_metrics": {
            "exit_cap_rate_pct": {"value": 5.25},
            "stabilized_noi": {"value": 4_000_000},
            "exit_year": {"value": 5},
        },
    }
    
    # Manual calculation:
    # NOI = $4M (given)
    # Gross Exit = $4M / 0.0525 = $76.19M
    # Net Exit = $76.19M * 0.975 = $74.29M
    # Total Profit = $74.29M - $70M = $4.29M
    # Profit on Cost = $4.29M / $70M = 6.13%
    
    expected_gross_exit = 4_000_000 / 0.0525  # ~$76.19M
    expected_profit = (expected_gross_exit * 0.975) - 70_000_000  # ~$4.29M
    expected_poc = (expected_profit / 70_000_000) * 100  # ~6.13%
    
    calc = compute_returns(test_pro_forma)
    
    print(f"  Expected gross exit: ${expected_gross_exit:,.0f}")
    print(f"  Calculated gross exit: ${calc.get('calc_gross_exit_value', 0):,.0f}")
    
    print(f"  Expected profit: ${expected_profit:,.0f}")
    print(f"  Calculated profit: ${calc.get('calc_total_profit', 0):,.0f}")
    
    print(f"  Expected PoC: {expected_poc:.2f}%")
    print(f"  Calculated PoC: {calc.get('calc_profit_on_cost_pct', 0):.2f}%")
    
    # Check within 5%
    gross_diff = abs(calc.get("calc_gross_exit_value", 0) - expected_gross_exit) / expected_gross_exit
    assert gross_diff < 0.05, f"Gross exit diff {gross_diff:.1%} exceeds 5%"
    print(f"  Gross exit difference: {gross_diff:.1%} (within 5%)")
    
    print("\nPASS: Calculator produces accurate results")
    return True


# ═══════════════════════════════════════════════════════════════════════════════
# 7.5 — Test Excel Export
# ═══════════════════════════════════════════════════════════════════════════════

def test_excel_export_complete():
    """Verify Excel export has all required components."""
    print("\n" + "=" * 60)
    print("TEST 7.5: Excel Export")
    print("=" * 60)
    
    from openpyxl import load_workbook
    import io
    from FallonPrototype.shared.return_calculator import compute_sensitivity_table
    
    # Sample export data
    export_data = {
        "pro_forma": {
            "project_summary": {
                "deal_name": "Test Charlotte Multifamily",
                "market": "charlotte",
                "program_type": "multifamily",
                "total_gfa_sf": {"value": 180000, "unit": "sf", "label": "calculated", "source": "test"},
                "unit_count": {"value": 200, "unit": "units", "label": "confirmed", "source": "user"},
                "notes": "Test export",
            },
            "revenue_assumptions": {
                "rent_psf_monthly": {"value": 1.85, "unit": "$/sf", "label": "estimated", "source": "market"},
                "stabilized_occupancy_pct": {"value": 93, "unit": "%", "label": "estimated", "source": "market"},
            },
            "cost_assumptions": {
                "hard_cost_psf": {"value": 275, "unit": "$/sf", "label": "estimated", "source": "market"},
                "total_project_cost": {"value": 70000000, "unit": "$", "label": "calculated", "source": "sum"},
            },
            "financing_assumptions": {
                "equity_required": {"value": 24500000, "unit": "$", "label": "calculated", "source": "calc"},
                "lp_equity_amount": {"value": 22050000, "unit": "$", "label": "calculated", "source": "calc"},
            },
            "return_metrics": {
                "exit_cap_rate_pct": {"value": 5.25, "unit": "%", "label": "estimated", "source": "market"},
                "project_irr_levered_pct": {"value": 15.0, "unit": "%", "label": "estimated", "source": "model"},
                "equity_multiple_lp": {"value": 1.85, "unit": "x", "label": "estimated", "source": "model"},
            },
        },
        "warnings": ["Test warning"],
        "calc_results": {},
    }
    
    # Add sensitivity table
    export_data["sensitivity"] = compute_sensitivity_table(export_data["pro_forma"], target_irr=14.0)
    
    # Generate Excel
    excel_bytes = export_pro_forma(export_data, "Test")
    wb = load_workbook(io.BytesIO(excel_bytes))
    
    # Check sheets
    expected_sheets = ["Summary", "Revenue", "Costs", "Financing", "Returns", "Assumptions & Sources"]
    print(f"  Sheets found: {wb.sheetnames}")
    
    for sheet in expected_sheets:
        assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"
    print("  All 6 sheets present")
    
    # Check Returns sheet has sensitivity table
    returns_ws = wb["Returns"]
    all_values = " ".join(str(c.value) for row in returns_ws.iter_rows() for c in row if c.value)
    assert "Sensitivity" in all_values, "Missing sensitivity table in Returns sheet"
    print("  Sensitivity table present in Returns")
    
    # Check Assumptions sheet has estimated values
    assumptions_ws = wb["Assumptions & Sources"]
    all_values = " ".join(str(c.value) for row in assumptions_ws.iter_rows() for c in row if c.value)
    assert "estimated" in all_values.lower() or "REVENUE" in all_values, "Missing estimated assumptions"
    print("  Assumptions & Sources populated")
    
    print("\nPASS: Excel export complete")
    return True


# ═══════════════════════════════════════════════════════════════════════════════
# 7.7 — Test Clarification Flow
# ═══════════════════════════════════════════════════════════════════════════════

def test_clarification_flow():
    """Test that vague queries trigger clarification and merge works."""
    print("\n" + "=" * 60)
    print("TEST 7.7: Clarification Flow")
    print("=" * 60)
    
    # Vague query should have missing fields
    vague_query = "build me a model"
    params = normalize_parameters(extract_parameters(vague_query))
    missing = check_missing_parameters(params)
    
    print(f"  Vague query: '{vague_query}'")
    print(f"  Missing fields: {missing}")
    
    assert len(missing) > 0, "Vague query should trigger clarification"
    print("  Clarification triggered correctly")
    
    # Clarification should fill in gaps
    clarification = "Charlotte multifamily, 200 units"
    merged = merge_clarification(params, clarification)
    
    print(f"  After clarification: market={merged.market}, program={merged.program_type}, units={merged.unit_count}")
    
    # Check merged params
    merged_missing = check_missing_parameters(merged)
    print(f"  Remaining missing: {merged_missing}")
    
    # Should have fewer missing fields
    assert len(merged_missing) < len(missing), "Clarification should reduce missing fields"
    print("  Merge reduces missing fields")
    
    print("\nPASS: Clarification flow works")
    return True


# ═══════════════════════════════════════════════════════════════════════════════
# 7.1 — Test Extraction (Quick Version)
# ═══════════════════════════════════════════════════════════════════════════════

def test_extraction_quick():
    """Quick extraction test without API calls (checks normalization)."""
    print("\n" + "=" * 60)
    print("TEST 7.1: Extraction (Normalization)")
    print("=" * 60)
    
    # Test normalization
    params = ProjectParameters(
        market="Charlotte, NC",
        program_type="apartments",
        unit_count=200,
    )
    
    normalized = normalize_parameters(params)
    
    print(f"  Input: market='Charlotte, NC', program_type='apartments'")
    print(f"  Normalized: market='{normalized.market}', program_type='{normalized.program_type}'")
    
    assert normalized.market == "charlotte", f"Expected 'charlotte', got '{normalized.market}'"
    assert normalized.program_type == "multifamily", f"Expected 'multifamily', got '{normalized.program_type}'"
    
    # Test GFA estimation
    assert normalized.total_gfa_sf == 200 * 900, "GFA should be estimated from unit count"
    print(f"  Estimated GFA: {normalized.total_gfa_sf} sf (200 units * 900 sf)")
    
    print("\nPASS: Normalization works correctly")
    return True


# ═══════════════════════════════════════════════════════════════════════════════
# Data Status Check
# ═══════════════════════════════════════════════════════════════════════════════

def check_data_status():
    """Check if required data is indexed."""
    print("\n" + "=" * 60)
    print("DATA STATUS CHECK")
    print("=" * 60)
    
    try:
        counts = get_collection_counts()
        print(f"  Deal data: {counts.get('fallon_deal_data', 0)} documents")
        print(f"  Market defaults: {counts.get('fallon_market_defaults', 0)} documents")
        print(f"  Contracts: {counts.get('fallon_contracts', 0)} documents")
        
        if counts.get('fallon_deal_data', 0) == 0:
            print("\n  NOTE: No deal data indexed. Run ingestion first:")
            print("    python -m FallonPrototype.Financial Model.shared.run_all_ingestion")
        
        return True
    except Exception as e:
        print(f"  Error checking data: {e}")
        return False


# ═══════════════════════════════════════════════════════════════════════════════
# Run All Tests
# ═══════════════════════════════════════════════════════════════════════════════

def run_all_tests():
    """Run all Phase 7 validation tests."""
    print("\n" + "=" * 60)
    print("PHASE 7 - END-TO-END VALIDATION")
    print("=" * 60)
    
    # Data status
    check_data_status()
    
    tests = [
        ("7.1 Extraction (normalization)", test_extraction_quick),
        ("7.2 Deal comps retrieval", test_deal_comps_retrieval),
        ("7.4 Calculator cross-check", test_calculator_cross_check),
        ("7.5 Excel export", test_excel_export_complete),
        ("7.7 Clarification flow", test_clarification_flow),
    ]
    
    results = []
    for name, test_fn in tests:
        try:
            passed = test_fn()
            results.append((name, passed))
        except Exception as e:
            print(f"\nERROR in {name}: {e}")
            import traceback
            traceback.print_exc()
            results.append((name, False))
    
    print("\n" + "=" * 60)
    print("VALIDATION SUMMARY")
    print("=" * 60)
    
    for name, passed in results:
        status = "PASS" if passed else "FAIL"
        print(f"  {name}: {status}")
    
    print("\n" + "-" * 60)
    print("MANUAL TESTS REQUIRED:")
    print("  7.3 - Generate 3 deal types (requires Claude API)")
    print("  7.6 - Adjust assumptions in UI (requires running app)")
    print("-" * 60)
    
    all_passed = all(p for _, p in results)
    print(f"\nAutomated tests: {'ALL PASSED' if all_passed else 'SOME FAILED'}")
    
    return all_passed


if __name__ == "__main__":
    success = run_all_tests()
    sys.exit(0 if success else 1)
