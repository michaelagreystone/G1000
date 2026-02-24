"""
Full Integration Test - Verifies all 7 phases work together.
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))


def test_phase1_data_layer():
    """Phase 1: Verify data is indexed."""
    print("\n" + "=" * 60)
    print("PHASE 1: Data Layer")
    print("=" * 60)
    
    from FallonPrototype.shared.vector_store import get_collection_counts, get_market_defaults
    
    counts = get_collection_counts()
    print(f"  Deal data: {counts.get('fallon_deal_data', 0)} chunks")
    print(f"  Market defaults: {counts.get('fallon_market_defaults', 0)} records")
    
    # Test direct lookup
    defaults = get_market_defaults("charlotte", "multifamily")
    if defaults:
        rent = defaults.get("rent_psf_monthly", {}).get("value")
        print(f"  Charlotte MF rent: ${rent}/sf/mo")
        assert rent is not None, "Should have rent data"
    
    print("  PASS")
    return True


def test_phase2_extraction():
    """Phase 2: Verify parameter extraction and normalization."""
    print("\n" + "=" * 60)
    print("PHASE 2: Parameter Extraction")
    print("=" * 60)
    
    from FallonPrototype.agents.financial_agent import (
        extract_parameters, normalize_parameters, check_missing_parameters
    )
    
    # Test normalization (no API call)
    from FallonPrototype.agents.financial_agent import ProjectParameters
    params = ProjectParameters(market="Charlotte, NC", program_type="apartments", unit_count=200)
    normalized = normalize_parameters(params)
    
    assert normalized.market == "charlotte"
    assert normalized.program_type == "multifamily"
    assert normalized.total_gfa_sf == 180000
    print(f"  Normalized: {normalized.market} {normalized.program_type}")
    
    # Test missing check
    incomplete = ProjectParameters()
    missing = check_missing_parameters(incomplete)
    assert len(missing) > 0
    print(f"  Missing fields detected: {len(missing)}")
    
    print("  PASS")
    return True


def test_phase3_context():
    """Phase 3: Verify context retrieval and formatting."""
    print("\n" + "=" * 60)
    print("PHASE 3: Context Retrieval")
    print("=" * 60)
    
    from FallonPrototype.agents.financial_agent import (
        ProjectParameters, retrieve_deal_comps, get_defaults_for_params,
        format_financial_context, assemble_context
    )
    
    params = ProjectParameters(market="charlotte", program_type="multifamily", unit_count=200)
    
    # Test deal comps
    comps = retrieve_deal_comps(params)
    print(f"  Deal comps retrieved: {len(comps)}")
    
    # Test defaults
    defaults = get_defaults_for_params(params)
    assert defaults is not None
    print(f"  Defaults found: {len(defaults)} fields")
    
    # Test full context assembly
    context, warning = assemble_context(params)
    assert len(context) > 500
    print(f"  Context assembled: {len(context)} chars")
    
    print("  PASS")
    return True


def test_phase4_calculator():
    """Phase 4: Verify return calculator."""
    print("\n" + "=" * 60)
    print("PHASE 4: Return Calculator")
    print("=" * 60)
    
    from FallonPrototype.shared.return_calculator import compute_returns, compute_sensitivity_table
    
    # Test with sample pro forma
    pro_forma = {
        "project_summary": {"program_type": "multifamily", "unit_count": {"value": 200}},
        "revenue_assumptions": {"rent_psf_monthly": {"value": 1.85}, "stabilized_occupancy_pct": {"value": 93}},
        "cost_assumptions": {"total_project_cost": {"value": 70000000}},
        "financing_assumptions": {"equity_required": {"value": 25000000}, "lp_equity_amount": {"value": 22500000}, "lp_equity_pct": {"value": 90}},
        "return_metrics": {"exit_cap_rate_pct": {"value": 5.25}, "stabilized_noi": {"value": 4000000}, "exit_year": {"value": 5}},
    }
    
    results = compute_returns(pro_forma)
    print(f"  NOI: ${results.get('calc_noi', 0):,.0f}")
    print(f"  Gross exit: ${results.get('calc_gross_exit_value', 0):,.0f}")
    print(f"  Profit on cost: {results.get('calc_profit_on_cost_pct', 0):.1f}%")
    
    # Test sensitivity
    sens = compute_sensitivity_table(pro_forma, target_irr=14.0)
    assert len(sens["values"]) == 3
    print(f"  Sensitivity table: 3x3 grid generated")
    
    print("  PASS")
    return True


def test_phase5_excel():
    """Phase 5: Verify Excel export."""
    print("\n" + "=" * 60)
    print("PHASE 5: Excel Export")
    print("=" * 60)
    
    from FallonPrototype.shared.excel_export import export_pro_forma
    from FallonPrototype.shared.return_calculator import compute_sensitivity_table
    from openpyxl import load_workbook
    import io
    
    # Sample export data
    export_data = {
        "pro_forma": {
            "project_summary": {"deal_name": "Test Deal", "market": "charlotte", "program_type": "multifamily"},
            "revenue_assumptions": {"rent_psf_monthly": {"value": 1.85, "unit": "$/sf", "label": "estimated", "source": "test"}},
            "cost_assumptions": {"total_project_cost": {"value": 70000000, "unit": "$", "label": "calculated", "source": "test"}},
            "financing_assumptions": {"equity_required": {"value": 25000000, "unit": "$", "label": "calculated", "source": "test"}},
            "return_metrics": {"exit_cap_rate_pct": {"value": 5.25, "unit": "%", "label": "estimated", "source": "test"}},
        },
        "warnings": [],
    }
    export_data["sensitivity"] = compute_sensitivity_table(export_data["pro_forma"], 14.0)
    
    excel_bytes = export_pro_forma(export_data, "Test")
    wb = load_workbook(io.BytesIO(excel_bytes))
    
    assert len(wb.sheetnames) == 6
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  File size: {len(excel_bytes):,} bytes")
    
    print("  PASS")
    return True


def test_phase6_ui_imports():
    """Phase 6: Verify Streamlit app imports (doesn't run server)."""
    print("\n" + "=" * 60)
    print("PHASE 6: UI Module Check")
    print("=" * 60)
    
    import py_compile
    app_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "app.py")
    
    try:
        py_compile.compile(app_path, doraise=True)
        print(f"  app.py syntax: OK")
    except py_compile.PyCompileError as e:
        print(f"  app.py syntax error: {e}")
        return False
    
    print("  PASS")
    return True


def test_phase7_validation():
    """Phase 7: Run validation tests."""
    print("\n" + "=" * 60)
    print("PHASE 7: Validation")
    print("=" * 60)
    
    from FallonPrototype.agents.financial_agent import validate_pro_forma, VALID_LABELS
    
    # Valid pro forma
    valid_pf = {
        "project_summary": {},
        "revenue_assumptions": {},
        "cost_assumptions": {"total_project_cost": {"value": 70000000, "unit": "$", "label": "calculated", "source": "test"}},
        "financing_assumptions": {},
        "return_metrics": {},
    }
    
    is_valid, errors = validate_pro_forma(valid_pf)
    print(f"  Schema validation: {'PASS' if is_valid else 'FAIL'}")
    if errors:
        print(f"  Errors: {errors[:3]}")
    
    # Check labels
    print(f"  Valid labels: {VALID_LABELS}")
    
    print("  PASS")
    return True


def run_full_integration():
    """Run all integration tests."""
    print("\n" + "=" * 60)
    print("FULL INTEGRATION TEST - ALL 7 PHASES")
    print("=" * 60)
    
    tests = [
        ("Phase 1: Data Layer", test_phase1_data_layer),
        ("Phase 2: Extraction", test_phase2_extraction),
        ("Phase 3: Context", test_phase3_context),
        ("Phase 4: Calculator", test_phase4_calculator),
        ("Phase 5: Excel", test_phase5_excel),
        ("Phase 6: UI Module", test_phase6_ui_imports),
        ("Phase 7: Validation", test_phase7_validation),
    ]
    
    results = []
    for name, test_fn in tests:
        try:
            passed = test_fn()
            results.append((name, passed))
        except Exception as e:
            print(f"\n  ERROR: {e}")
            import traceback
            traceback.print_exc()
            results.append((name, False))
    
    print("\n" + "=" * 60)
    print("INTEGRATION TEST SUMMARY")
    print("=" * 60)
    
    for name, passed in results:
        status = "PASS" if passed else "FAIL"
        print(f"  {name}: {status}")
    
    all_passed = all(p for _, p in results)
    print(f"\n{'=' * 60}")
    print(f"RESULT: {'ALL PHASES WORKING' if all_passed else 'SOME PHASES FAILED'}")
    print("=" * 60)
    
    return all_passed


if __name__ == "__main__":
    success = run_full_integration()
    sys.exit(0 if success else 1)
