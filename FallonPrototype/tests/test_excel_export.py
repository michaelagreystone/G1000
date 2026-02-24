"""
Tests for Phase 5 — Excel Export.
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from FallonPrototype.shared.excel_export import (
    export_pro_forma,
    get_suggested_filename,
)
from FallonPrototype.shared.return_calculator import compute_sensitivity_table

# Sample export data for testing
SAMPLE_EXPORT_DATA = {
    "pro_forma": {
        "project_summary": {
            "deal_name": "Charlotte Multifamily 200 Units",
            "market": "charlotte",
            "program_type": "multifamily",
            "total_gfa_sf": {"value": 180000, "unit": "sf", "label": "calculated", "source": "unit_count * 900"},
            "unit_count": {"value": 200, "unit": "units", "label": "confirmed", "source": "user-provided"},
            "rentable_sf": None,
            "construction_start": {"value": "Q3 2026", "unit": "", "label": "confirmed", "source": "user-provided"},
            "construction_duration_months": {"value": 18, "unit": "months", "label": "estimated", "source": "Charlotte market"},
            "total_keys": None,
            "notes": "Test project for Excel export validation",
        },
        "revenue_assumptions": {
            "rent_psf_monthly": {"value": 1.85, "unit": "$/sf/month", "label": "estimated", "source": "Charlotte broker data Q4 2025"},
            "rent_psf_annual_nnn": None,
            "adr": None,
            "stabilized_occupancy_pct": {"value": 93, "unit": "%", "label": "estimated", "source": "Charlotte market data"},
            "lease_up_months": {"value": 18, "unit": "months", "label": "estimated", "source": "Charlotte market standard"},
            "annual_rent_growth_pct": {"value": 3.0, "unit": "%", "label": "estimated", "source": "Charlotte forecast"},
            "other_income_per_unit_monthly": {"value": 125, "unit": "$/unit/month", "label": "estimated", "source": "Charlotte market"},
        },
        "cost_assumptions": {
            "land_cost_total": {"value": 8000000, "unit": "$", "label": "estimated", "source": "Charlotte land market"},
            "hard_cost_psf": {"value": 275, "unit": "$/sf", "label": "estimated", "source": "Charlotte construction Q4 2025"},
            "hard_cost_total": {"value": 49500000, "unit": "$", "label": "calculated", "source": "hard_cost_psf * total_gfa_sf"},
            "soft_cost_pct_of_hard": {"value": 22, "unit": "%", "label": "estimated", "source": "Industry standard"},
            "soft_cost_total": {"value": 10890000, "unit": "$", "label": "calculated", "source": "hard_cost * soft_cost_pct"},
            "developer_fee_pct": {"value": 4, "unit": "%", "label": "estimated", "source": "Fallon standard"},
            "developer_fee_total": {"value": 2735600, "unit": "$", "label": "calculated", "source": "total_cost * dev_fee_pct"},
            "contingency_pct": {"value": 7, "unit": "%", "label": "estimated", "source": "Standard range"},
            "contingency_total": {"value": 3465000, "unit": "$", "label": "calculated", "source": "hard_cost * contingency_pct"},
            "total_project_cost": {"value": 74590600, "unit": "$", "label": "calculated", "source": "sum of all costs"},
        },
        "financing_assumptions": {
            "construction_loan_ltc_pct": {"value": 65, "unit": "%", "label": "estimated", "source": "Market standard"},
            "construction_loan_amount": {"value": 48483890, "unit": "$", "label": "calculated", "source": "ltc * total_cost"},
            "construction_loan_rate_pct": {"value": 7.5, "unit": "%", "label": "estimated", "source": "SOFR + spread"},
            "carry_cost_total": {"value": 5453438, "unit": "$", "label": "calculated", "source": "loan * rate * term/12"},
            "equity_required": {"value": 26106710, "unit": "$", "label": "calculated", "source": "total_cost - loan"},
            "lp_equity_pct": {"value": 90, "unit": "%", "label": "estimated", "source": "Fallon standard"},
            "lp_equity_amount": {"value": 23496039, "unit": "$", "label": "calculated", "source": "equity * lp_pct"},
            "gp_equity_pct": {"value": 10, "unit": "%", "label": "estimated", "source": "Fallon standard"},
            "gp_equity_amount": {"value": 2610671, "unit": "$", "label": "calculated", "source": "equity * gp_pct"},
        },
        "return_metrics": {
            "exit_cap_rate_pct": {"value": 5.25, "unit": "%", "label": "estimated", "source": "Charlotte cap rate Q4 2025"},
            "exit_year": {"value": 5, "unit": "years", "label": "estimated", "source": "Fallon standard hold"},
            "stabilized_noi": {"value": 4200000, "unit": "$", "label": "calculated", "source": "revenue - expenses"},
            "gross_exit_value": {"value": 80000000, "unit": "$", "label": "calculated", "source": "noi / cap_rate"},
            "net_exit_value": {"value": 78000000, "unit": "$", "label": "calculated", "source": "gross - sale_costs"},
            "total_profit": {"value": 3409400, "unit": "$", "label": "calculated", "source": "net_exit - total_cost"},
            "profit_on_cost_pct": {"value": 4.57, "unit": "%", "label": "calculated", "source": "profit / total_cost"},
            "development_spread_bps": {"value": 38, "unit": "bps", "label": "calculated", "source": "(yield - cap) * 10000"},
            "project_irr_levered_pct": {"value": 15.5, "unit": "%", "label": "estimated", "source": "DCF estimate"},
            "equity_multiple_lp": {"value": 1.85, "unit": "x", "label": "estimated", "source": "LP return calc"},
            "lp_irr_pct": {"value": 14.2, "unit": "%", "label": "estimated", "source": "LP DCF estimate"},
        },
    },
    "warnings": [
        "Note: Construction costs may vary based on final design.",
        "Exit cap rate assumption should be verified with broker.",
    ],
    "calc_results": {
        "calc_noi": 4200000,
        "calc_profit_on_cost_pct": 4.57,
    },
}


def test_export_pro_forma_generates_bytes():
    """Test that export_pro_forma returns bytes."""
    print("\n" + "=" * 60)
    print("TEST: export_pro_forma() generates bytes")
    print("=" * 60)
    
    # Add sensitivity table
    export_data = SAMPLE_EXPORT_DATA.copy()
    export_data["sensitivity"] = compute_sensitivity_table(
        export_data["pro_forma"],
        target_irr=14.0
    )
    
    result = export_pro_forma(export_data, "Test Pro Forma")
    
    assert isinstance(result, bytes)
    assert len(result) > 0
    
    # Check it starts with Excel magic bytes (PK for zip format)
    assert result[:2] == b'PK'
    
    print(f"  Generated {len(result):,} bytes")
    print("  Starts with Excel magic bytes (PK)")
    
    print("\nPASS: Excel export generates valid bytes")
    return True


def test_export_has_all_sheets():
    """Test that the exported workbook has all required sheets."""
    print("\n" + "=" * 60)
    print("TEST: Workbook has all 6 sheets")
    print("=" * 60)
    
    from openpyxl import load_workbook
    import io
    
    export_data = SAMPLE_EXPORT_DATA.copy()
    export_data["sensitivity"] = compute_sensitivity_table(
        export_data["pro_forma"],
        target_irr=14.0
    )
    
    excel_bytes = export_pro_forma(export_data, "Test")
    wb = load_workbook(io.BytesIO(excel_bytes))
    
    expected_sheets = ["Summary", "Revenue", "Costs", "Financing", "Returns", "Assumptions & Sources"]
    
    print(f"  Found sheets: {wb.sheetnames}")
    
    for sheet in expected_sheets:
        assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"
        print(f"    - {sheet}: OK")
    
    print("\nPASS: All 6 sheets present")
    return True


def test_summary_sheet_content():
    """Test that Summary sheet has key metrics."""
    print("\n" + "=" * 60)
    print("TEST: Summary sheet has key metrics")
    print("=" * 60)
    
    from openpyxl import load_workbook
    import io
    
    export_data = SAMPLE_EXPORT_DATA.copy()
    export_data["sensitivity"] = compute_sensitivity_table(
        export_data["pro_forma"],
        target_irr=14.0
    )
    
    excel_bytes = export_pro_forma(export_data, "Test")
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Summary"]
    
    # Check deal name in first cell
    assert "Charlotte Multifamily" in str(ws.cell(row=1, column=1).value)
    print("  - Deal name present")
    
    # Check for metric labels
    all_values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
    content = " ".join(all_values)
    
    assert "IRR" in content or "Levered" in content
    assert "Multiple" in content or "Equity" in content
    print("  - Key metrics present")
    
    print("\nPASS: Summary sheet has key content")
    return True


def test_sensitivity_table_in_returns():
    """Test that Returns sheet includes sensitivity table."""
    print("\n" + "=" * 60)
    print("TEST: Returns sheet has sensitivity table")
    print("=" * 60)
    
    from openpyxl import load_workbook
    import io
    
    export_data = SAMPLE_EXPORT_DATA.copy()
    export_data["sensitivity"] = compute_sensitivity_table(
        export_data["pro_forma"],
        target_irr=14.0
    )
    
    excel_bytes = export_pro_forma(export_data, "Test")
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Returns"]
    
    # Look for sensitivity table title
    all_values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
    content = " ".join(all_values)
    
    assert "Sensitivity" in content
    assert "-10%" in content or "Base" in content
    print("  - Sensitivity table title found")
    print("  - Column headers present")
    
    print("\nPASS: Returns sheet has sensitivity table")
    return True


def test_get_suggested_filename():
    """Test filename suggestion."""
    print("\n" + "=" * 60)
    print("TEST: get_suggested_filename()")
    print("=" * 60)
    
    filename = get_suggested_filename(SAMPLE_EXPORT_DATA)
    
    assert filename.endswith(".xlsx")
    assert "Charlotte" in filename
    print(f"  Suggested filename: {filename}")
    
    print("\nPASS: Filename suggestion works")
    return True


def test_save_sample_export():
    """Generate and save a sample Excel file for manual inspection."""
    print("\n" + "=" * 60)
    print("TEST: Save sample export file")
    print("=" * 60)
    
    export_data = SAMPLE_EXPORT_DATA.copy()
    export_data["sensitivity"] = compute_sensitivity_table(
        export_data["pro_forma"],
        target_irr=14.0
    )
    
    excel_bytes = export_pro_forma(export_data, "Test")
    
    # Save to test output directory
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "sample_pro_forma_export.xlsx")
    
    with open(output_path, "wb") as f:
        f.write(excel_bytes)
    
    print(f"  Saved to: {output_path}")
    print(f"  File size: {len(excel_bytes):,} bytes")
    
    print("\nPASS: Sample file saved for inspection")
    return True


def run_all_tests():
    """Run all Phase 5 tests."""
    print("\n" + "=" * 60)
    print("PHASE 5 - EXCEL EXPORT TESTS")
    print("=" * 60)
    
    tests = [
        ("export_generates_bytes", test_export_pro_forma_generates_bytes),
        ("has_all_sheets", test_export_has_all_sheets),
        ("summary_content", test_summary_sheet_content),
        ("sensitivity_table", test_sensitivity_table_in_returns),
        ("suggested_filename", test_get_suggested_filename),
        ("save_sample", test_save_sample_export),
    ]
    
    results = []
    for name, test_fn in tests:
        try:
            passed = test_fn()
            results.append((name, passed))
        except Exception as e:
            print(f"\nERROR in {name}: {e}")
            import traceback
            traceback.print_exc()
            results.append((name, False))
    
    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)
    
    for name, passed in results:
        status = "PASS" if passed else "FAIL"
        print(f"  {name}: {status}")
    
    all_passed = all(p for _, p in results)
    print(f"\nOverall: {'ALL TESTS PASSED' if all_passed else 'SOME TESTS FAILED'}")
    
    return all_passed


if __name__ == "__main__":
    success = run_all_tests()
    sys.exit(0 if success else 1)
